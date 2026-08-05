#!/usr/bin/env python3
"""Import Assembly OEE workbooks into the Line Side Board SQL Server schema."""

from __future__ import annotations

import datetime as dt
import json
import re
import sys
from pathlib import Path


def emit(summary: dict) -> None:
    print(json.dumps(summary, default=str, separators=(",", ":")))


try:
    import openpyxl
    import pyodbc
except Exception as exc:  # pragma: no cover - depends on deployment machine
    emit({"stations": 0, "oee_entries": 0, "issues": 0, "errors": [f"Missing Python dependency: {exc}"]})
    raise SystemExit(2)


ALIASES = {
    "date": {"date", "entrydate", "recordeddate", "productiondate"},
    "shift": {"shift", "shiftnum", "shiftnumber"},
    "cell": {"cell", "station", "stationname", "resource", "resourceid", "workcenter"},
    "part_number": {"part", "partnumber", "partno", "pn"},
    "tool_dt": {"tooldt", "toolingdt", "toolingdowntime", "tooldowntimemin"},
    "top_tool_issue": {"toptoolissue", "toolingissue", "toolissue"},
    "maint_dt": {"maintdt", "maintenancedt", "maintenancedowntime", "maintdowntimemin"},
    "top_maint_issue": {"topmaintissue", "maintenanceissue", "maintissue"},
    "prod_dt": {"proddt", "productiondt", "productiondowntime", "proddowntimemin"},
    "top_prod_issue": {"topprodissue", "productionissue", "prodissue"},
    "parts": {"parts", "partsreported", "reportedparts", "actual", "actualqty", "totalparts"},
    "plan": {"plan", "planned", "plannedqty", "targetparts"},
    "target_cycle": {"targetcycletime", "targetcyclesec", "tct"},
    "actual_cycle": {"actualcycletime", "actualcyclesec", "act"},
    "scrap": {"scrap", "scrapqty", "scrapcount"},
    "rework": {"rework", "reworkqty", "reworkcount"},
    "availability": {"availability", "availabilitypct", "uptime", "uptimepct"},
    "performance": {"performance", "performancepct", "efficiency", "efficiencypct", "eff"},
    "quality": {"quality", "qualitypct", "fpy", "fpypct"},
    "oee": {"oee", "oeepct", "oeepercent"},
    "operator": {"operator", "operatorname", "employee"},
}

ALIAS_LOOKUP = {alias: canonical for canonical, aliases in ALIASES.items() for alias in aliases}


def normalise(value: object) -> str:
    return re.sub(r"[^a-z0-9]", "", str(value or "").strip().lower())


def number(value: object, default: float = 0.0) -> float:
    if value is None or value == "":
        return default
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "").replace("%", "")
    try:
        return float(text)
    except ValueError:
        return default


def integer(value: object, default: int = 0) -> int:
    return int(round(number(value, float(default))))


def percent(value: object, default: float = 0.0) -> float:
    result = number(value, default)
    if 0 < result <= 1:
        result *= 100
    return max(0.0, min(100.0, result))


def date_value(value: object) -> dt.date:
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    if value:
        text = str(value).strip()
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%d/%m/%Y"):
            try:
                return dt.datetime.strptime(text, fmt).date()
            except ValueError:
                pass
    return dt.date.today()


def find_header(sheet) -> tuple[int | None, dict[int, str]]:
    best_row = None
    best_map: dict[int, str] = {}
    for row_number, row in enumerate(sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 30), values_only=True), 1):
        mapped = {}
        for index, value in enumerate(row):
            canonical = ALIAS_LOOKUP.get(normalise(value))
            if canonical:
                mapped[index] = canonical
        if len(mapped) > len(best_map):
            best_row, best_map = row_number, mapped
    return best_row, best_map


def row_dict(values: tuple, mapping: dict[int, str]) -> dict[str, object]:
    return {canonical: values[index] if index < len(values) else None for index, canonical in mapping.items()}


def ensure_schema(cursor) -> None:
    cursor.execute("""
    IF OBJECT_ID('dbo.oee_entries', 'U') IS NULL
    CREATE TABLE dbo.oee_entries (
        id INT IDENTITY PRIMARY KEY, entry_date DATE, shift INT, cell VARCHAR(100),
        part_number VARCHAR(100), tool_dt_min INT DEFAULT 0, top_tool_issue VARCHAR(500),
        maint_dt_min INT DEFAULT 0, top_maint_issue VARCHAR(500), prod_dt_min INT DEFAULT 0,
        top_prod_issue VARCHAR(500), parts_reported INT DEFAULT 0,
        target_cycle_sec DECIMAL(10,2), actual_cycle_sec DECIMAL(10,2), scrap INT DEFAULT 0,
        rework INT DEFAULT 0, availability_pct DECIMAL(6,2), performance_pct DECIMAL(6,2),
        quality_pct DECIMAL(6,2), oee_pct DECIMAL(6,2), created_at DATETIME DEFAULT GETDATE()
    )
    """)
    cursor.execute("""
    IF OBJECT_ID('dbo.kpi_summary', 'U') IS NULL
    CREATE TABLE dbo.kpi_summary (
        resource_id VARCHAR(100), shift INT, actual INT, plan INT, efficiency_pct DECIMAL(6,2),
        fpy_pct DECIMAL(6,2), avg_cycle_sec DECIMAL(10,2), hours_worked DECIMAL(6,2),
        open_issues INT, jph_target INT, operator_name VARCHAR(100), recorded_date DATE,
        CONSTRAINT pk_kpi_summary PRIMARY KEY (resource_id, shift, recorded_date)
    )
    """)


def insert_oee(cursor, record: dict[str, object]) -> None:
    parts = integer(record.get("parts"))
    scrap = integer(record.get("scrap"))
    rework = integer(record.get("rework"))
    tool_dt = integer(record.get("tool_dt"))
    maint_dt = integer(record.get("maint_dt"))
    prod_dt = integer(record.get("prod_dt"))
    availability = percent(record.get("availability"), max(0, (480 - tool_dt - maint_dt - prod_dt) / 480 * 100))
    target_cycle = number(record.get("target_cycle"))
    actual_cycle = number(record.get("actual_cycle"))
    performance = percent(record.get("performance"), (target_cycle / actual_cycle * 100) if actual_cycle else 0)
    quality = percent(record.get("quality"), ((parts - scrap - rework) / parts * 100) if parts else 0)
    oee = percent(record.get("oee"), availability * performance * quality / 10000)
    cursor.execute("""
        INSERT INTO dbo.oee_entries (
            entry_date, shift, cell, part_number, tool_dt_min, top_tool_issue,
            maint_dt_min, top_maint_issue, prod_dt_min, top_prod_issue,
            parts_reported, target_cycle_sec, actual_cycle_sec, scrap, rework,
            availability_pct, performance_pct, quality_pct, oee_pct
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, date_value(record.get("date")), max(1, integer(record.get("shift"), 2)), str(record.get("cell") or "").strip(),
        str(record.get("part_number") or "").strip(), tool_dt, str(record.get("top_tool_issue") or "").strip(),
        maint_dt, str(record.get("top_maint_issue") or "").strip(), prod_dt, str(record.get("top_prod_issue") or "").strip(),
        parts, target_cycle, actual_cycle, scrap, rework, availability, performance, quality, oee)


def merge_kpi(cursor, record: dict[str, object]) -> None:
    cell = str(record.get("cell") or "").strip()
    if not cell:
        return
    entry_date = date_value(record.get("date"))
    shift = max(1, integer(record.get("shift"), 2))
    actual = integer(record.get("parts"))
    efficiency = percent(record.get("performance"))
    plan = integer(record.get("plan"))
    if plan <= 0 and efficiency > 0:
        plan = int(round(actual / (efficiency / 100)))
    quality = percent(record.get("quality"), 100)
    cycle = number(record.get("actual_cycle"))
    operator = str(record.get("operator") or "").strip()
    cursor.execute("""
        MERGE dbo.kpi_summary AS target
        USING (SELECT ? AS resource_id, ? AS shift, ? AS recorded_date) AS source
        ON target.resource_id = source.resource_id AND target.shift = source.shift AND target.recorded_date = source.recorded_date
        WHEN MATCHED THEN UPDATE SET actual=?, plan=?, efficiency_pct=?, fpy_pct=?, avg_cycle_sec=?, operator_name=?
        WHEN NOT MATCHED THEN INSERT (resource_id, shift, recorded_date, actual, plan, efficiency_pct, fpy_pct, avg_cycle_sec, hours_worked, open_issues, jph_target, operator_name)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, 8, 0, 0, ?);
    """, cell, shift, entry_date, actual, plan, efficiency, quality, cycle, operator,
        cell, shift, entry_date, actual, plan, efficiency, quality, cycle, operator)


def main() -> int:
    if len(sys.argv) != 3:
        emit({"stations": 0, "oee_entries": 0, "issues": 0, "errors": ["Usage: import_excel.py workbook.xlsx connection-string"]})
        return 2
    workbook_path = Path(sys.argv[1])
    if not workbook_path.is_file() or workbook_path.suffix.lower() != ".xlsx":
        emit({"stations": 0, "oee_entries": 0, "issues": 0, "errors": ["A valid .xlsx workbook is required"]})
        return 2

    errors: list[str] = []
    stations: set[str] = set()
    imported = 0
    connection = None
    try:
        workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
        selected = [sheet for sheet in workbook.worksheets if sheet.title in {"OEE Data", "Station OEE", "Operator Data"} or "oee" in sheet.title.lower()]
        if not selected:
            raise ValueError("No OEE-compatible sheets were found")

        connection = pyodbc.connect(sys.argv[2], timeout=15, autocommit=False)
        cursor = connection.cursor()
        ensure_schema(cursor)
        for sheet in selected:
            header_row, mapping = find_header(sheet)
            if header_row is None or "cell" not in mapping.values():
                errors.append(f"{sheet.title}: no station/cell header mapping found")
                continue
            if "oee" not in mapping.values() and "parts" not in mapping.values():
                errors.append(f"{sheet.title}: no OEE or parts column found")
                continue
            for values in sheet.iter_rows(min_row=header_row + 1, values_only=True):
                record = row_dict(values, mapping)
                cell = str(record.get("cell") or "").strip()
                if not cell or not any(value not in (None, "") for value in record.values()):
                    continue
                insert_oee(cursor, record)
                merge_kpi(cursor, record)
                stations.add(cell)
                imported += 1
        connection.commit()
    except Exception as exc:
        if connection is not None:
            connection.rollback()
        errors.append(str(exc))
    finally:
        if connection is not None:
            connection.close()

    emit({"stations": len(stations), "oee_entries": imported, "issues": 0, "errors": errors})
    return 0 if imported or not errors else 1


if __name__ == "__main__":
    raise SystemExit(main())
